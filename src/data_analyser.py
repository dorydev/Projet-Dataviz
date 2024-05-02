import openpyxl
from openpyxl.utils import column_index_from_string
import re


# cette classe est tres utile pour lire une matrice delimitee par un mot d entree premiere ligne 
# et pour un mot de sortie pour la derniere ligne 
# on pourra peut etre rajouter une logique permettant de gerer l entete du fichier
# mais ce n est pas a l ordre du jour 
class DataReader1:
    def __init__(self, file_path , start_row_key_word , end_row_keyword  ):
        self.file_path = file_path
        self.matrice = []

        self.main_matrice_start_row = 0
        self.main_matrice_end_row = 0
        self.start_row_key_word =  start_row_key_word 
        self.end_row_keyword  =  end_row_keyword  
        self.workbook =  openpyxl.load_workbook(self.file_path)
        self.determine_start_Row_and_end_Row( start_row_key_word , end_row_keyword   )
        # self.start_main_row = start_main_row


    # permet de déterminer le debut d une ligne et la fin d une ligne du fichier excel  
    def determine_start_Row_and_end_Row(self, key_word_start_row, key_word_end_row, column_search=1):
        try:
            # print(" column search : " + str(column_search))
            sheet = self.workbook.active
            #print("Sheet")
            #print(str(sheet))
            start_row = 1
            end_row = 1
            found_start_row = False
            found_end_row = False

            while start_row < sheet.max_row and not found_start_row  :
                cell_value = sheet.cell(row=start_row, column=column_search).value
                if re.search(str(key_word_start_row), str(cell_value), re.IGNORECASE):
                    found_start_row = True
                else:
                    start_row += 1

            end_row =  start_row + 1  

            while end_row < sheet.max_row and not found_end_row :
                cell_value = sheet.cell(row=end_row, column=column_search).value
                if re.search(str(key_word_end_row), str(cell_value), re.IGNORECASE):
                    found_end_row = True
                else:
                    end_row += 1

            self.main_matrice_start_row = start_row
            self.main_matrice_end_row = end_row - 1

            print("Numero de la ligne où ça commence : ", start_row)
            print("Numero de la ligne où ça finit : ", end_row)

        except Exception as e:
            print("Problème lors de la détermination de la première et dernière ligne de la matrice :", e)

    def return_main_first_row_header(self ):
        # Fetch the row from the worksheet
        row = self.workbook.active [self.main_matrice_start_row]
        # Convert each cell in the row to a string and return as a list
        return [str(cell.value) if cell.value is not None else "" for cell in row]

    # permet de charger une colonne en particulier 
    def load_column(self, column_number, start_row , limit_row):
        try:
            # self.workbook = openpyxl.load_self.workbook(self.file_path)
            sheet = self.workbook.active
            column_data = []
            for row in range(start_row,  limit_row - start_row):
                cell_value = sheet.cell(row=row, column=column_number).value
                column_data.append(cell_value)
            return column_data
        except FileNotFoundError:
            print("Le fichier spécifié est introuvable.")
            return None
        except Exception as e:
            print("Une erreur s'est produite lors du chargement de la colonne :", e)
            return None

    # permet de charger toute la colonne par le nom 
    def load_all_columns_by_name(self, row_number  , end_row_number  , pattern):
        try:
            all_columns_data = []
            sheet = self.workbook.active
            # print(" VALEUR DE SHEET " + str( sheet.cell( row=self.main_matrice_start_row , column=2).value  ) )
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_number, column=col).value
                if cell_value and re.search(pattern, str(cell_value), re.IGNORECASE):
                    column_data = self.load_column(col, row_number , end_row_number)
                    #print(column_data)
                    all_columns_data.append(column_data)

            #print(" JUSQUE LA ")
            # Affichage de la matrice ligne par ligne
            #print(" affichage ligne par ligne de  " + str( all_columns_data ))
            # for row_data in zip(*all_columns_data):
                #print(row_data) 
            return all_columns_data
        except FileNotFoundError:
            #print("Le fichier spécifié est introuvable.")
            return None
        except Exception as e:
            print("Une erreur s'est produite lors du chargement de toutes les colonnes par nom :", e)
            return None 

    # permet de charger toutes les colonnes a partir d un pattern  (par ex type obj_12 pour avoir toutes les infos d une UE ) 
    def load_all_columns_by_name_from_the_current_line(self,  pattern):
        return self.load_all_columns_by_name( self.main_matrice_start_row , self.main_matrice_end_row , pattern )
    
    def load_matrix_from_current_line_name( self , listePattern  ):
        matrix = [ ]
        for pattern in listePattern: 
            for subColonne in  self.load_all_columns_by_name_from_the_current_line(pattern):
                matrix.append(subColonne)

        # Affichage de la matrice ligne par ligne
        # for row_data in zip(*matrix):
            #print(row_data) 

    # obtenir le num de ligne et de colonne à partir du titre de la colonne et de la valeur recherchee 
    def getRowIndexValue( self ,  title_column , value_that_we_are_looking_for  ):
        indexRow = self.main_matrice_start_row
        column = self.load_all_columns_by_name(self.main_matrice_start_row , self.main_matrice_end_row , title_column)
        for items in column:
            if items and  re.search(value_that_we_are_looking_for   , str(items), re.IGNORECASE):
                break 
            else:
                indexRow += 1 
        indexRow += 1 
        #print(" Le numero de la ligne de la valeur " + str(indexRow))
        return indexRow

    # renvoit un index en lettre de la colonne courante 
    def getColIndexLetterValue(self , titleColumn):
        try:
            sheet = self.workbook.active
            ligne = sheet[self.main_matrice_start_row]
            for items in ligne:
                # print(str(items.value) )
                if items and re.search(titleColumn , str(items.value), re.IGNORECASE):
                    return items.column_letter
            return ""
        except FileNotFoundError:
            print("Le fichier spécifié est introuvable.")
            return None
        except Exception as e:
            print("Une erreur s'est produite lors du chargement de toutes les colonnes par nom :", e)
            return None 

    # 
    def get_corresponding_value_of_another_column( self ,  title_column_1 , value_column_1 ,  title_column_2  ) :
        ligne_referent =  self.getRowIndexValue( title_column_1 , value_column_1  )
        lettre_colonne_cherchee = self.getColIndexLetterValue( title_column_2)
        sheet = self.workbook.active
        # Accéder à la cellule et lire sa valeur
        #print(" VALEUR RECHERCHEE " + str( lettre_colonne_cherchee   ) + str(ligne_referent) )
        #print( sheet[lettre_colonne_cherchee  + str(ligne_referent)].value)
        return sheet[lettre_colonne_cherchee  + str(ligne_referent)].value

    def get_column_letter(self , column_index):
        """ Convertit un index de colonne numérique en lettre de colonne. """
        string = ""
        while column_index > 0:
            column_index, remainder = divmod(column_index - 1, 26)
            string = chr(65 + remainder) + string
        return string
     
    def loading_full_line( self , row_number  ):
        sheet = self.workbook.active 
        row = sheet[ row_number  ]
        row_values = [ cell.value for cell in row ]
        return row_values 
    
    def loading_header_matrice_and_one_tuple( self , title_column_referent , value_referent ):
        header_line = self.loading_full_line( self.main_matrice_start_row )
        row_valued_index = self.getRowIndexValue( title_column_referent ,  value_referent  ) 
        row_valued = self.loading_full_line( row_valued_index  )
        dictionary_returned = { key : value  for key , value in zip( header_line  ,  row_valued  )}

        # for key , value in dictionary_returned.items():
            #print(f"Key: {key}, Value: {value}")

        return  dictionary_returned 


class MatrixHandler:
    def returnIndexedListFromHeadingMatrice( matrix ):
        column_dict =  { row[0] : row[1::] for row in matrix } 
        return  column_dict 

    def eliminateNonNumberCharacters(data_dict):
        # Iterate over each key-value pair in the dictionary
        for key, values in data_dict.items():
            # Process each item in the list
            new_values = []
            for value in values:
                if isinstance(value, str)  :  # Check if the item is a string
                    # Remove all non-digit characters from the string
                    cleaned_value = re.sub(r'\D', '', value)
                    new_values.append(cleaned_value)
                else:
                    new_values.append(value)
            # Update the dictionary with the cleaned list
            data_dict[key] = new_values

        return data_dict

    def printDictionnary( example_dict ):
        # Find the maximum length of the lists to ensure all columns are fully displayed
        max_length = max(len(values) for values in example_dict.values())

        # Print the keys as column headers
        print(' | '.join(example_dict.keys()))

        # Iterate through each index up to the maximum length
        for i in range(max_length):
            # For each key, get the value at index i if it exists, otherwise print a placeholder
            row = [str(example_dict[key][i]) if i < len(example_dict[key]) else '---' for key in example_dict]
            print(' | '.join(row))
    def print_dict_keys(input_dict):
        # Print all keys in the dictionary
        for key in input_dict.keys():
            print(key)

    def find_corresponding_value(data_dict, search_key, search_value, target_key):
        """
        Search for a value in a dictionary under a specified key and retrieve the corresponding value from another key.

        Parameters:
        - data_dict (dict): The dictionary containing the data.
        - search_key (str): The key under which to search for the value.
        - search_value (any): The value to search for.
        - target_key (str): The key from which to retrieve the corresponding value.

        Returns:
        - The corresponding value from the target key or None if the search value is not found.
        """
        try:
            # Find the index of the search value under the search key
            index = data_dict[search_key].index(search_value)
            # Retrieve the corresponding value from the target key
            return data_dict[target_key][index]
        except (ValueError, KeyError):
            # Return None if the search value is not found or keys are incorrect
            return None

    def find_index_in_list(value_list, value):
        """
        Retrieve the index of a value in a list.

        Parameters:
        - value_list (list): The list to search through.
        - value (any): The value to find the index of.

        Returns:
        - int: The index of the value in the list, or None if the value is not found.
       1"""
        index = 0
        for items in value_list:
            if items == value:
                return index 
            index += 1 
        return 1
    
    def extract_numbers( strings):
        # Regular expression to find numbers in the given format
        pattern = r"Obj(\d+)"
        # List to store all the numbers found
        numbers = []
        
        # Iterate over each string in the list
        for string in strings:
            # Search for the pattern and extract the number
            matches = re.findall(pattern, string)
            if matches:
                # Convert the found numbers to integers and add to the list
                numbers.extend([int(num) for num in matches])

        MatrixHandler.remove_duplicates_in_place(numbers)
            
        # Return the list of numbers
        return numbers 

    def remove_duplicates_in_place( input_list):
        seen = set()
        index = 0
        while index < len(input_list):
            item = input_list[index]
            if item in seen:
                input_list.pop(index)
            else:
                seen.add(item)
                index += 1



    # calcul le max de Objn dans le fichier excel 
    def find_max_n( strings):
        # Regular expression to find numbers followed by any characters
        pattern = r"Obj(\d+)_.*"
        
        # List to store all the numbers found
        numbers = []
        
        # Iterate over each string in the list
        for string in strings:
            # Search for the pattern and extract the number
            match = re.search(pattern, string)
            if match:
                # Convert the found number to an integer and add to the list
                number = int(match.group(1))
                numbers.append(number)
        
        # Calculate the maximum number in the list, if the list is not empty
        if numbers:
            return max(numbers)
        else:
            return 0 # Return None if no numbers were found


class UniteEnseignement:
    def __init__( self , fichierUEDetaille , numeroUE , DataReaderUnique ):
        self.fichierUEDetaille   =  fichierUEDetaille 
        self.dataReaderUE =  DataReaderUnique   
        # on stock es listes avec cle valeur , exemple   Obj2_Result : { ........ } , Obj2_truc : { .......  }
        self.nomObj =  "Obj" + str(numeroUE) + "_"
        all_columns_ue_temp = self.dataReaderUE.load_all_columns_by_name_from_the_current_line("Obj" + str(numeroUE) + "_")
        self.all_columns_ue =  MatrixHandler.returnIndexedListFromHeadingMatrice( all_columns_ue_temp )

        #print(" ---- print dict keys  --------  ")
        # MatrixHandler.printDictionnary(self.all_columns_ue)

        # on stock les numeros etudiants  
        all_student_temp =  self.dataReaderUE.load_all_columns_by_name_from_the_current_line("Etud_Numér")   
        self.all_student =  MatrixHandler.returnIndexedListFromHeadingMatrice( all_student_temp )
        self.all_student =  MatrixHandler.eliminateNonNumberCharacters( self.all_student)
        # MatrixHandler.printDictionnary( self.all_student  )
        # test retriev 
        self.getStudentGradeForThisUe("583902")

    def get_libelle_UE(self  ):
        for items in  self.all_columns_ue[self.nomObj + "Libellé"]:
            if items is not None and items is not "":
                return items 
        return ""

    def get_student_list(self):
        return  self.all_student 

    def get_a_specific_field_from_a_student( self , num_Etudiant , name_of_the_field ): 
        index = MatrixHandler.find_index_in_list( self.all_student["Etud_Numér"] ,   num_Etudiant  )
        # print(" On retrieve la note de " + str( num_UE ) + " qui vaut : " +  self.all_columns_ue[ self.nomObj + name_of_the_field ][index]  )
        return  self.all_columns_ue[ self.nomObj +  name_of_the_field  ][index]

    # def get_grade_of_the_student_in_this_UE( numeroEtudiant ,   )
    def getStudentGradeForThisUe( self , num_Etudiant ):
        # print("  ---- grade -----  " )
        # print( self.get_a_specific_field_from_a_student(  num_Etudiant , "Note_Ado/20") )
        return self.get_a_specific_field_from_a_student(  num_Etudiant , "Note_Ado/20")

    def getStudentAdmissabilityInThisUe( self , num_Etudiant ):
        return  self.get_a_specific_field_from_a_student(  num_Etudiant , "Résultat")
        # self.all_columns_ue =  self.dataReaderUE.load_all_columns_by_name_from_the_current_line("Obj" + str(num_UE))

class UEManager: 
    def __init__(self , fichierUEDetaille  ):

        self.fichierUEDetaille  =  fichierUEDetaille  
        self.pre_data_reader = DataReader1( fichierUEDetaille , "Etud_Numér"  , "XX_FIN_PV_XX" )
        listeEntete  = self.pre_data_reader.return_main_first_row_header()  

        listeIndex = MatrixHandler.extract_numbers( listeEntete  )

        # On commence a initialiser la liste des UE 
        # contient la liste des UE 
        self.listeUEClass =   []
        print(" ------- liste des UE -----------  ")

        i  = 0 
        for numUe in listeIndex:
            #print( numUe)
            self.listeUEClass.append( UniteEnseignement( self.fichierUEDetaille ,  numUe , self.pre_data_reader ) )
            #print(self.listeUEClass[i].get_libelle_UE())
            i += 1 

        # on stock les numeros etudiants  
        all_student_temp =  self.pre_data_reader.load_all_columns_by_name_from_the_current_line("Etud_Numér")   
        self.all_student =  MatrixHandler.returnIndexedListFromHeadingMatrice( all_student_temp )
        self.all_student =  MatrixHandler.eliminateNonNumberCharacters( self.all_student)


    def getListeOfUE(self ): 
        return self.listeUEClass 
    
    def getStudentListe(self ):
        return  self.all_student
    
    def getBulletinOfAStudent(self ,  num_etudiant ):
        listOfGrades = []
        listOFUELibelle = []
        for ue in self.listeUEClass:
            # Try to convert the string to a float
            try:
                listOfGrades.append( ue.getStudentGradeForThisUe(num_etudiant))
                listOFUELibelle.append( ue.get_libelle_UE() )
                #print("Converted number:",  float(  ue.getStudentGradeForThisUe(num_etudiant)  )   )
            except ValueError:
                print("The string is not a valid number.")
        return dict(zip(  listOFUELibelle ,  listOfGrades   ) )

class  GlobalResult:
    def __init__(self ,   dataReader ):
        self.dataReader =  dataReader 
        all_columns_ue_temp = self.dataReader.load_all_columns_by_name_from_the_current_line("Result_")
        self.all_columns_ue =  MatrixHandler.returnIndexedListFromHeadingMatrice( all_columns_ue_temp )
        self.liste_of_num_students =  self.dataReader.load_all_columns_by_name_from_the_current_line("Etud_Numér")
        self.liste_of_num_students =  MatrixHandler.returnIndexedListFromHeadingMatrice( self.liste_of_num_students )

    def get_grades(self ):
        print(" ------------------ GET GRADES -----------------------")
        #for items in  self.all_columns_ue["Result_Note_Ado/20"]:
            #print(items)
        return self.all_columns_ue["Result_Note_Ado/20"]
    
    def get_students(self ):
        print(" ------------------ GET GRADES -----------------------")
        #for items in  self.liste_of_num_students["Etud_Numér"]:
            #print(items)
        return self.liste_of_num_students

    

# classe constituant une structure de donnee de base pour un etudiant  
class EtudiantSearcherInDetailedFile:
    def __init__(self, fichierIdentitéEtudiant , fichierUEDetaille , numeroEtudiantStringified  ):
        self.fichierIdentitéEtudiant =   fichierIdentitéEtudiant 
        self.fichierUEDetaille   =  fichierUEDetaille   

        self.data_reader_fichier_identité_Etudiant = DataReader1( fichierIdentitéEtudiant ,  "Rentrée" , "")
        self.data_reader_fichierUEDetaillee = DataReader1(  fichierUEDetaille   ,  "Etud_Numér" , "XX_FIN_PV_XX")

        self.numeroEtudiant =  numeroEtudiantStringified  
        self.attributs =  self.data_reader_fichier_identité_Etudiant.loading_header_matrice_and_one_tuple("Numéro d'étudiant" ,str(numeroEtudiantStringified) )

    # pourquoi pas ...  
    def getNumberOfCredits():
        return 

class L1PCResultsExtractor:
    def __init__(self, file_path , start_row_key_word , end_row_keyword  ):
        self.filepath = file_path 
        self.start_row_key_word 
        self.end_row_keyword =  end_row_keyword 


# Exemple d'utilisationimport openpyxl
from openpyxl.utils import column_index_from_string
import re


# cette classe est tres utile pour lire une matrice delimitee par un mot d entree premiere ligne 
# et pour un mot de sortie pour la derniere ligne 
# on pourra peut etre rajouter une logique permettant de gerer l entete du fichier
# mais ce n est pas a l ordre du jour 
class DataReader1:
    def __init__(self, file_path , start_row_key_word , end_row_keyword  ):
        self.file_path = file_path
        self.matrice = []

        self.main_matrice_start_row = 0
        self.main_matrice_end_row = 0
        self.start_row_key_word =  start_row_key_word 
        self.end_row_keyword  =  end_row_keyword  
        self.workbook =  openpyxl.load_workbook(self.file_path)
        self.determine_start_Row_and_end_Row( start_row_key_word , end_row_keyword   )
        # self.start_main_row = start_main_row


    # permet de déterminer le debut d une ligne et la fin d une ligne du fichier excel  
    def determine_start_Row_and_end_Row(self, key_word_start_row, key_word_end_row, column_search=1):
        try:
            # print(" column search : " + str(column_search))
            sheet = self.workbook.active
            #print("Sheet")
            #print(str(sheet))
            start_row = 1
            end_row = 1
            found_start_row = False
            found_end_row = False

            while start_row < sheet.max_row and not found_start_row  :
                cell_value = sheet.cell(row=start_row, column=column_search).value
                if re.search(str(key_word_start_row), str(cell_value), re.IGNORECASE):
                    found_start_row = True
                else:
                    start_row += 1

            end_row =  start_row + 1  

            while end_row < sheet.max_row and not found_end_row :
                cell_value = sheet.cell(row=end_row, column=column_search).value
                if re.search(str(key_word_end_row), str(cell_value), re.IGNORECASE):
                    found_end_row = True
                else:
                    end_row += 1

            self.main_matrice_start_row = start_row
            self.main_matrice_end_row = end_row - 1

            print("Numero de la ligne où ça commence : ", start_row)
            print("Numero de la ligne où ça finit : ", end_row)

        except Exception as e:
            print("Problème lors de la détermination de la première et dernière ligne de la matrice :", e)

    def return_main_first_row_header(self ):
        # Fetch the row from the worksheet
        row = self.workbook.active [self.main_matrice_start_row]
        # Convert each cell in the row to a string and return as a list
        return [str(cell.value) if cell.value is not None else "" for cell in row]

    # permet de charger une colonne en particulier 
    def load_column(self, column_number, start_row , limit_row):
        try:
            # self.workbook = openpyxl.load_self.workbook(self.file_path)
            sheet = self.workbook.active
            column_data = []
            for row in range(start_row,  limit_row - start_row):
                cell_value = sheet.cell(row=row, column=column_number).value
                column_data.append(cell_value)
            return column_data
        except FileNotFoundError:
            print("Le fichier spécifié est introuvable.")
            return None
        except Exception as e:
            print("Une erreur s'est produite lors du chargement de la colonne :", e)
            return None

    # permet de charger toute la colonne par le nom 
    def load_all_columns_by_name(self, row_number  , end_row_number  , pattern):
        try:
            all_columns_data = []
            sheet = self.workbook.active
            # print(" VALEUR DE SHEET " + str( sheet.cell( row=self.main_matrice_start_row , column=2).value  ) )
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_number, column=col).value
                if cell_value and re.search(pattern, str(cell_value), re.IGNORECASE):
                    column_data = self.load_column(col, row_number , end_row_number)
                    #print(column_data)
                    all_columns_data.append(column_data)

            #print(" JUSQUE LA ")
            # Affichage de la matrice ligne par ligne
            #print(" affichage ligne par ligne de  " + str( all_columns_data ))
            # for row_data in zip(*all_columns_data):
                #print(row_data) 
            return all_columns_data
        except FileNotFoundError:
            #print("Le fichier spécifié est introuvable.")
            return None
        except Exception as e:
            print("Une erreur s'est produite lors du chargement de toutes les colonnes par nom :", e)
            return None 

    # permet de charger toutes les colonnes a partir d un pattern  (par ex type obj_12 pour avoir toutes les infos d une UE ) 
    def load_all_columns_by_name_from_the_current_line(self,  pattern):
        return self.load_all_columns_by_name( self.main_matrice_start_row , self.main_matrice_end_row , pattern )
    
    def load_matrix_from_current_line_name( self , listePattern  ):
        matrix = [ ]
        for pattern in listePattern: 
            for subColonne in  self.load_all_columns_by_name_from_the_current_line(pattern):
                matrix.append(subColonne)

        # Affichage de la matrice ligne par ligne
        # for row_data in zip(*matrix):
            #print(row_data) 

    # obtenir le num de ligne et de colonne à partir du titre de la colonne et de la valeur recherchee 
    def getRowIndexValue( self ,  title_column , value_that_we_are_looking_for  ):
        indexRow = self.main_matrice_start_row
        column = self.load_all_columns_by_name(self.main_matrice_start_row , self.main_matrice_end_row , title_column)
        for items in column:
            if items and  re.search(value_that_we_are_looking_for   , str(items), re.IGNORECASE):
                break 
            else:
                indexRow += 1 
        indexRow += 1 
        #print(" Le numero de la ligne de la valeur " + str(indexRow))
        return indexRow

    # renvoit un index en lettre de la colonne courante 
    def getColIndexLetterValue(self , titleColumn):
        try:
            sheet = self.workbook.active
            ligne = sheet[self.main_matrice_start_row]
            for items in ligne:
                # print(str(items.value) )
                if items and re.search(titleColumn , str(items.value), re.IGNORECASE):
                    return items.column_letter
            return ""
        except FileNotFoundError:
            print("Le fichier spécifié est introuvable.")
            return None
        except Exception as e:
            print("Une erreur s'est produite lors du chargement de toutes les colonnes par nom :", e)
            return None 

    # 
    def get_corresponding_value_of_another_column( self ,  title_column_1 , value_column_1 ,  title_column_2  ) :
        ligne_referent =  self.getRowIndexValue( title_column_1 , value_column_1  )
        lettre_colonne_cherchee = self.getColIndexLetterValue( title_column_2)
        sheet = self.workbook.active
        # Accéder à la cellule et lire sa valeur
        #print(" VALEUR RECHERCHEE " + str( lettre_colonne_cherchee   ) + str(ligne_referent) )
        #print( sheet[lettre_colonne_cherchee  + str(ligne_referent)].value)
        return sheet[lettre_colonne_cherchee  + str(ligne_referent)].value

    def get_column_letter(self , column_index):
        """ Convertit un index de colonne numérique en lettre de colonne. """
        string = ""
        while column_index > 0:
            column_index, remainder = divmod(column_index - 1, 26)
            string = chr(65 + remainder) + string
        return string
     
    def loading_full_line( self , row_number  ):
        sheet = self.workbook.active 
        row = sheet[ row_number  ]
        row_values = [ cell.value for cell in row ]
        return row_values 
    
    def loading_header_matrice_and_one_tuple( self , title_column_referent , value_referent ):
        header_line = self.loading_full_line( self.main_matrice_start_row )
        row_valued_index = self.getRowIndexValue( title_column_referent ,  value_referent  ) 
        row_valued = self.loading_full_line( row_valued_index  )
        dictionary_returned = { key : value  for key , value in zip( header_line  ,  row_valued  )}

        # for key , value in dictionary_returned.items():
            #print(f"Key: {key}, Value: {value}")

        return  dictionary_returned 


class MatrixHandler:
    def returnIndexedListFromHeadingMatrice( matrix ):
        column_dict =  { row[0] : row[1::] for row in matrix } 
        return  column_dict 

    def eliminateNonNumberCharacters(data_dict):
        # Iterate over each key-value pair in the dictionary
        for key, values in data_dict.items():
            # Process each item in the list
            new_values = []
            for value in values:
                if isinstance(value, str)  :  # Check if the item is a string
                    # Remove all non-digit characters from the string
                    cleaned_value = re.sub(r'\D', '', value)
                    new_values.append(cleaned_value)
                else:
                    new_values.append(value)
            # Update the dictionary with the cleaned list
            data_dict[key] = new_values

        return data_dict

    def printDictionnary( example_dict ):
        # Find the maximum length of the lists to ensure all columns are fully displayed
        max_length = max(len(values) for values in example_dict.values())

        # Print the keys as column headers
        print(' | '.join(example_dict.keys()))

        # Iterate through each index up to the maximum length
        for i in range(max_length):
            # For each key, get the value at index i if it exists, otherwise print a placeholder
            row = [str(example_dict[key][i]) if i < len(example_dict[key]) else '---' for key in example_dict]
            print(' | '.join(row))
    def print_dict_keys(input_dict):
        # Print all keys in the dictionary
        for key in input_dict.keys():
            print(key)

    def find_corresponding_value(data_dict, search_key, search_value, target_key):
        """
        Search for a value in a dictionary under a specified key and retrieve the corresponding value from another key.

        Parameters:
        - data_dict (dict): The dictionary containing the data.
        - search_key (str): The key under which to search for the value.
        - search_value (any): The value to search for.
        - target_key (str): The key from which to retrieve the corresponding value.

        Returns:
        - The corresponding value from the target key or None if the search value is not found.
        """
        try:
            # Find the index of the search value under the search key
            index = data_dict[search_key].index(search_value)
            # Retrieve the corresponding value from the target key
            return data_dict[target_key][index]
        except (ValueError, KeyError):
            # Return None if the search value is not found or keys are incorrect
            return None

    def find_index_in_list(value_list, value):
        """
        Retrieve the index of a value in a list.

        Parameters:
        - value_list (list): The list to search through.
        - value (any): The value to find the index of.

        Returns:
        - int: The index of the value in the list, or None if the value is not found.
       1"""
        index = 0
        for items in value_list:
            if items == value:
                return index 
            index += 1 
        return 1
    
    def extract_numbers( strings):
        # Regular expression to find numbers in the given format
        pattern = r"Obj(\d+)"
        # List to store all the numbers found
        numbers = []
        
        # Iterate over each string in the list
        for string in strings:
            # Search for the pattern and extract the number
            matches = re.findall(pattern, string)
            if matches:
                # Convert the found numbers to integers and add to the list
                numbers.extend([int(num) for num in matches])

        MatrixHandler.remove_duplicates_in_place(numbers)
            
        # Return the list of numbers
        return numbers 

    def remove_duplicates_in_place( input_list):
        seen = set()
        index = 0
        while index < len(input_list):
            item = input_list[index]
            if item in seen:
                input_list.pop(index)
            else:
                seen.add(item)
                index += 1



    # calcul le max de Objn dans le fichier excel 
    def find_max_n( strings):
        # Regular expression to find numbers followed by any characters
        pattern = r"Obj(\d+)_.*"
        
        # List to store all the numbers found
        numbers = []
        
        # Iterate over each string in the list
        for string in strings:
            # Search for the pattern and extract the number
            match = re.search(pattern, string)
            if match:
                # Convert the found number to an integer and add to the list
                number = int(match.group(1))
                numbers.append(number)
        
        # Calculate the maximum number in the list, if the list is not empty
        if numbers:
            return max(numbers)
        else:
            return 0 # Return None if no numbers were found


class UniteEnseignement:
    def __init__( self , fichierUEDetaille , numeroUE , DataReaderUnique ):
        self.fichierUEDetaille   =  fichierUEDetaille 
        self.dataReaderUE =  DataReaderUnique   
        # on stock es listes avec cle valeur , exemple   Obj2_Result : { ........ } , Obj2_truc : { .......  }
        self.nomObj =  "Obj" + str(numeroUE) + "_"
        all_columns_ue_temp = self.dataReaderUE.load_all_columns_by_name_from_the_current_line("Obj" + str(numeroUE) + "_")
        self.all_columns_ue =  MatrixHandler.returnIndexedListFromHeadingMatrice( all_columns_ue_temp )

        #print(" ---- print dict keys  --------  ")
        # MatrixHandler.printDictionnary(self.all_columns_ue)

        # on stock les numeros etudiants  
        all_student_temp =  self.dataReaderUE.load_all_columns_by_name_from_the_current_line("Etud_Numér")   
        self.all_student =  MatrixHandler.returnIndexedListFromHeadingMatrice( all_student_temp )
        self.all_student =  MatrixHandler.eliminateNonNumberCharacters( self.all_student)
        # MatrixHandler.printDictionnary( self.all_student  )
        # test retriev 
        self.getStudentGradeForThisUe("583902")

    def get_libelle_UE(self  ):
        for items in  self.all_columns_ue[self.nomObj + "Libellé"]:
            if items is not None and items is not "":
                return items 
        return ""

    def get_student_list(self):
        return  self.all_student 

    def get_a_specific_field_from_a_student( self , num_Etudiant , name_of_the_field ): 
        index = MatrixHandler.find_index_in_list( self.all_student["Etud_Numér"] ,   num_Etudiant  )
        # print(" On retrieve la note de " + str( num_UE ) + " qui vaut : " +  self.all_columns_ue[ self.nomObj + name_of_the_field ][index]  )
        return  self.all_columns_ue[ self.nomObj +  name_of_the_field  ][index]

    # def get_grade_of_the_student_in_this_UE( numeroEtudiant ,   )
    def getStudentGradeForThisUe( self , num_Etudiant ):
        # print("  ---- grade -----  " )
        # print( self.get_a_specific_field_from_a_student(  num_Etudiant , "Note_Ado/20") )
        return self.get_a_specific_field_from_a_student(  num_Etudiant , "Note_Ado/20")

    def getStudentAdmissabilityInThisUe( self , num_Etudiant ):
        return  self.get_a_specific_field_from_a_student(  num_Etudiant , "Résultat")
        # self.all_columns_ue =  self.dataReaderUE.load_all_columns_by_name_from_the_current_line("Obj" + str(num_UE))

class UEManager: 
    def __init__(self , fichierUEDetaille  ):

        self.fichierUEDetaille  =  fichierUEDetaille  
        self.pre_data_reader = DataReader1( fichierUEDetaille , "Etud_Numér"  , "XX_FIN_PV_XX" )
        listeEntete  = self.pre_data_reader.return_main_first_row_header()  

        listeIndex = MatrixHandler.extract_numbers( listeEntete  )

        # On commence a initialiser la liste des UE 
        # contient la liste des UE 
        self.listeUEClass =   []
        print(" ------- liste des UE -----------  ")

        i  = 0 
        for numUe in listeIndex:
            print( numUe)
            self.listeUEClass.append( UniteEnseignement( self.fichierUEDetaille ,  numUe , self.pre_data_reader ) )
            print(self.listeUEClass[i].get_libelle_UE())
            i += 1 

        # on stock les numeros etudiants  
        all_student_temp =  self.pre_data_reader.load_all_columns_by_name_from_the_current_line("Etud_Numér")   
        self.all_student =  MatrixHandler.returnIndexedListFromHeadingMatrice( all_student_temp )
        self.all_student =  MatrixHandler.eliminateNonNumberCharacters( self.all_student)


    def getListeOfUE(self ): 
        return self.listeUEClass 
    
    def getStudentListe(self ):
        return  self.all_student
    
    def getBulletinOfAStudent(self ,  num_etudiant ):
        listOfGrades = []
        listOFUELibelle = []
        for ue in self.listeUEClass:
            # Try to convert the string to a float
            try:
                listOfGrades.append(  float(  ue.getStudentGradeForThisUe(num_etudiant)  )  )
                listOFUELibelle.append( ue.get_libelle_UE() )
                print("Converted number:",  float(  ue.getStudentGradeForThisUe(num_etudiant)  )   )
            except ValueError:
                print("The string is not a valid number.")
        return dict(zip(  listOFUELibelle ,  listOfGrades   ) )

class  GlobalResult:
    def __init__(self ,   dataReader ):
        self.dataReader =  dataReader 
        all_columns_ue_temp = self.dataReader.load_all_columns_by_name_from_the_current_line("Result_")
        self.all_columns_ue =  MatrixHandler.returnIndexedListFromHeadingMatrice( all_columns_ue_temp )
        self.liste_of_num_students =  self.dataReader.load_all_columns_by_name_from_the_current_line("Etud_Numér")
        self.liste_of_num_students =  MatrixHandler.returnIndexedListFromHeadingMatrice( self.liste_of_num_students )

    def get_grades(self ):
        print(" ------------------ GET GRADES -----------------------")
        #for items in  self.all_columns_ue["Result_Note_Ado/20"]:
            #print(items)
        return self.all_columns_ue["Result_Note_Ado/20"]
    
    def get_students(self ):
        print(" ------------------ GET GRADES -----------------------")
        #for items in  self.liste_of_num_students["Etud_Numér"]:
        #    print(items)
        return self.liste_of_num_students

    

# classe constituant une structure de donnee de base pour un etudiant  
class EtudiantSearcherInDetailedFile:
    def __init__(self, fichierIdentitéEtudiant , fichierUEDetaille , numeroEtudiantStringified  ):
        self.fichierIdentitéEtudiant =   fichierIdentitéEtudiant 
        self.fichierUEDetaille   =  fichierUEDetaille   

        self.data_reader_fichier_identité_Etudiant = DataReader1( fichierIdentitéEtudiant ,  "Rentrée" , "")
        self.data_reader_fichierUEDetaillee = DataReader1(  fichierUEDetaille   ,  "Etud_Numér" , "XX_FIN_PV_XX")

        self.numeroEtudiant =  numeroEtudiantStringified  
        self.attributs =  self.data_reader_fichier_identité_Etudiant.loading_header_matrice_and_one_tuple("Numéro d'étudiant" ,str(numeroEtudiantStringified) )

    # pourquoi pas ...  
    def getNumberOfCredits():
        return 

class L1PCResultsExtractor:
    def __init__(self, file_path , start_row_key_word , end_row_keyword  ):
        self.filepath = file_path 
        self.start_row_key_word 
        self.end_row_keyword =  end_row_keyword 


# Exemple d'utilisation
if __name__ == "__main__":
    file_path = "PV_L1_PC.xlsx"  # Spécifiez le chemin vers votre fichier Excel ici
    file_path_2  = "ODE_Extraction_AOA_L1_PC_2021.xlsx" 

    # data_reader = DataReader1(file_path, "Etud_" , "XX_FIN_PV_XX")
    # data_reader.getRowIndexValue( "Etud_Numér" , "583903" )
    # data_reader.getColIndexValue("Naissance")
    # data_reader.get_corresponding_value_of_another_column("Etud_Numér" ,  "583903" , "Result_Crédits" )

    # data_reader_etudiant  = DataReader1(file_path_2, "Rentrée" , None)
    # data_reader_etudiant.loading_header_matrice_and_one_tuple("Numéro d'étudiant" , "511581")

    # uniteEnseignement2 = UniteEnseignement(file_path , 2)


    # ue_manager_first = UEManager( file_path  )
    # pre_data_reader = DataReader1( file_path , "Etud_Numér"  , "XX_FIN_PV_XX" )
    # global_result =  GlobalResult(pre_data_reader)
    # global_result.get_grades()
    # global_result.get_students()


    #  loading_header_matrice_and_one_tuple
    # row_number = 11  # Numéro de la ligne à partir de laquelle charger les données
    # column_name = "Result_Note"  # Nom de la colonne à rechercher
    # column_data = data_reader.load_all_columns_by_name_from_the_current_line( pattern=column_name    )
    # column_data = data_reader.

if __name__ == "__main__":
    file_path = "PV_L1_PC.xlsx"  # Spécifiez le chemin vers votre fichier Excel ici
    file_path_2  = "ODE_Extraction_AOA_L1_PC_2021.xlsx" 

    # data_reader = DataReader1(file_path, "Etud_" , "XX_FIN_PV_XX")
    # data_reader.getRowIndexValue( "Etud_Numér" , "583903" )
    # data_reader.getColIndexValue("Naissance")
    # data_reader.get_corresponding_value_of_another_column("Etud_Numér" ,  "583903" , "Result_Crédits" )

    # data_reader_etudiant  = DataReader1(file_path_2, "Rentrée" , None)
    # data_reader_etudiant.loading_header_matrice_and_one_tuple("Numéro d'étudiant" , "511581")

    # uniteEnseignement2 = UniteEnseignement(file_path , 2)


    # ue_manager_first = UEManager( file_path  )
    # pre_data_reader = DataReader1( file_path , "Etud_Numér"  , "XX_FIN_PV_XX" )
    # global_result =  GlobalResult(pre_data_reader)
    # global_result.get_grades()
    # global_result.get_students()


    #  loading_header_matrice_and_one_tuple
    # row_number = 11  # Numéro de la ligne à partir de laquelle charger les données
    # column_name = "Result_Note"  # Nom de la colonne à rechercher
    # column_data = data_reader.load_all_columns_by_name_from_the_current_line( pattern=column_name    )
    # column_data = data_reader.
